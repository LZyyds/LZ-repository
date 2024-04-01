import openpyxl
import pandas as pd
import os


def v_lookup(current_workbook_name, current_worksheet_name, target_workbook_name, target_worksheet_name, column_id,
             start_row, end_row):
    # 打开当前工作簿和目标工作簿
    current_wb = openpyxl.load_workbook(current_workbook_name + '.xlsx')
    # target_wb = openpyxl.load_workbook(target_workbook_name + '.xlsx')

    # 通过工作表名称获取工作表对象
    current_ws = current_wb[current_worksheet_name]
    # target_ws = target_wb[target_worksheet_name]

    # 读取目标工作子表的数据到pandas DataFrame
    # target_df = pd.DataFrame(target_ws.values, columns=[cell.value for cell in target_ws[1]])
    target_df = pd.read_excel(target_workbook_name + '.xlsx', sheet_name=target_worksheet_name)
    # print(target_df)
    for i in range(start_row, end_row, 18):
        # 读取目标商品ID、商品名称
        target_id = current_ws[f"A{i}"].value
        target_name = current_ws[f"A{i - 1}"].value

        # 如果目标商品id在副表中找不到
        if target_id not in target_df['商品ID'].tolist():
            for n in range(i, i + 17):
                current_ws[f"{column_id}{n}"].value = 0
            print(f"********** {target_id} -- 未找到 -- [{target_name}] ***************")
        else:
            # 定义VLOOKUP的列名
            vlookup_columns = {
                f"{column_id}{i}": "支付金额",
                f"{column_id}{i + 1}": "自然流销售额",
                f"{column_id}{i + 3}": "商品访客数",
                f"{column_id}{i + 4}": "自然流访客",
                f"{column_id}{i + 9}": "付费流量加购收藏人数",
                f"{column_id}{i + 11}": "花费",
                f"{column_id}{i + 12}": "支付买家数",
                f"{column_id}{i + 13}": "支付新买家数",
                f"{column_id}{i + 15}": "付费流买家数",
                f"{column_id}{i + 16}": "新客支付金额"
            }

            # 在当前工作表中执行VLOOKUP操作
            for cell, target_column_name in vlookup_columns.items():
                current_ws[str(cell)].value = \
                    target_df.loc[target_df['商品ID'] == target_id, target_column_name].values[0]
                # print(current_ws[str(cell)].value)

            # 进行减法和除法运算
            current_ws[f"{column_id}{i + 2}"].value = float(current_ws[f"{column_id}{i}"].value) - float(
                current_ws[f"{column_id}{i + 1}"].value)
            current_ws[f"{column_id}{i + 5}"].value = float(current_ws[f"{column_id}{i + 3}"].value) - float(
                current_ws[f"{column_id}{i + 4}"].value)
            current_ws[f"{column_id}{i + 14}"].value = float(current_ws[f"{column_id}{i + 12}"].value) - float(
                current_ws[f"{column_id}{i + 15}"].value)

            # 分母为零的情况
            if current_ws[f"{column_id}{i + 4}"].value == 0.0:
                current_ws[f"{column_id}{i + 6}"].value = 0
            else:
                current_ws[f"{column_id}{i + 6}"].value = (float(current_ws[f"{column_id}{i + 14}"].value) / float(
                    current_ws[f"{column_id}{i + 4}"].value))
            if current_ws[f"{column_id}{i + 5}"].value == 0.0:
                current_ws[f"{column_id}{i + 7}"].value = 0
                current_ws[f"{column_id}{i + 8}"].value = 0
            else:
                current_ws[f"{column_id}{i + 7}"].value = (float(current_ws[f"{column_id}{i + 15}"].value) / float(
                    current_ws[f"{column_id}{i + 5}"].value))
                current_ws[f"{column_id}{i + 8}"].value = (float(current_ws[f"{column_id}{i + 9}"].value) / float(
                    current_ws[f"{column_id}{i + 5}"].value))
            if current_ws[f"{column_id}{i + 11}"].value == 0.0:
                current_ws[f"{column_id}{i + 10}"].value = 0
            else:
                current_ws[f"{column_id}{i + 10}"].value = (float(current_ws[f"{column_id}{i + 2}"].value) / float(
                    current_ws[f"{column_id}{i + 11}"].value))
            print(f"********** {target_worksheet_name} -- [{target_name}]已填完 ***************")

    print("\n\n")
    # 根据C列单元格清空后续列
    column_values = list(current_ws["C"])

    for idx, cell in enumerate(column_values):
        if cell.value is None:
            if not any(cell.coordinate in merged_range for merged_range in current_ws.merged_cells.ranges):
                current_ws[f"{column_id}{idx + 1}"].value = None
            # current_ws[f"{column_id}{idx + 1}"].value = None
    # 保存并关闭工作簿
    current_wb.save(current_workbook_name + '.xlsx')
    current_wb.close()
    # target_wb.close()


def v_lookup_with_tax(current_workbook_name, current_worksheet_name, target_workbook_name, target_worksheet_name,
                      column_id, start_row, end_row):
    # 打开当前工作簿和目标工作簿
    current_wb = openpyxl.load_workbook(current_workbook_name + '.xlsx')
    # target_wb = openpyxl.load_workbook(target_workbook_name + '.xlsx')

    # 通过工作表名称获取工作表对象
    current_ws = current_wb[current_worksheet_name]
    # target_ws = target_wb[target_worksheet_name]

    # 读取目标工作子表的数据到pandas DataFrame
    # target_df = pd.DataFrame(target_ws.values, columns=[cell.value for cell in target_ws[1]])
    target_df = pd.read_excel(target_workbook_name + '.xlsx', sheet_name=target_worksheet_name)
    # print(target_df)
    for i in range(start_row, end_row, 19):
        # 读取目标商品ID
        target_id = current_ws[f"A{i}"].value
        target_name = current_ws[f"A{i - 1}"].value

        if target_id not in target_df['商品ID'].tolist():
            for n in range(i, i + 18):
                current_ws[f"{column_id}{n}"].value = 0
            print(f"********** {target_id} -- 未找到 -- [{target_name}] ***************")
        else:
            # 定义VLOOKUP的列名
            vlookup_columns = {
                f"{column_id}{i}": "支付金额",
                f"{column_id}{i + 1 + 1}": "自然流销售额",
                f"{column_id}{i + 3 + 1}": "商品访客数",
                f"{column_id}{i + 4 + 1}": "自然流访客",
                f"{column_id}{i + 9 + 1}": "付费流量加购收藏人数",
                f"{column_id}{i + 11 + 1}": "花费",
                f"{column_id}{i + 12 + 1}": "支付买家数",
                f"{column_id}{i + 13 + 1}": "支付新买家数",
                f"{column_id}{i + 15 + 1}": "付费流买家数",
                f"{column_id}{i + 16 + 1}": "新客支付金额"
            }

            # 在当前工作表中执行VLOOKUP操作
            for cell, target_column_name in vlookup_columns.items():
                current_ws[str(cell)].value = \
                    target_df.loc[target_df['商品ID'] == target_id, target_column_name].values[0]
                # print(current_ws[str(cell)].value)

            # 进行减法和除法运算
            current_ws[f"{column_id}{i + 1}"].value = float(current_ws[f"{column_id}{i}"].value) * (1 - 0.091)
            current_ws[f"{column_id}{i + 2 + 1}"].value = float(current_ws[f"{column_id}{i}"].value) - float(
                current_ws[f"{column_id}{i + 1 + 1}"].value)
            current_ws[f"{column_id}{i + 5 + 1}"].value = float(current_ws[f"{column_id}{i + 3 + 1}"].value) - float(
                current_ws[f"{column_id}{i + 4 + 1}"].value)
            current_ws[f"{column_id}{i + 14 + 1}"].value = float(current_ws[f"{column_id}{i + 12 + 1}"].value) - float(
                current_ws[f"{column_id}{i + 15 + 1}"].value)

            # 分母为零的情况
            if current_ws[f"{column_id}{i + 4 + 1}"].value == 0.0:
                current_ws[f"{column_id}{i + 6 + 1}"].value = 0
            else:
                current_ws[f"{column_id}{i + 6 + 1}"].value = (
                        float(current_ws[f"{column_id}{i + 14 + 1}"].value) / float(
                    current_ws[f"{column_id}{i + 4 + 1}"].value))
            if current_ws[f"{column_id}{i + 5 + 1}"].value == 0.0:
                current_ws[f"{column_id}{i + 7 + 1}"].value = 0
                current_ws[f"{column_id}{i + 8 + 1}"].value = 0
            else:
                current_ws[f"{column_id}{i + 7 + 1}"].value = (
                        float(current_ws[f"{column_id}{i + 15 + 1}"].value) / float(
                    current_ws[f"{column_id}{i + 5 + 1}"].value))
                current_ws[f"{column_id}{i + 8 + 1}"].value = (
                        float(current_ws[f"{column_id}{i + 9 + 1}"].value) / float(
                    current_ws[f"{column_id}{i + 5 + 1}"].value))
            if current_ws[f"{column_id}{i + 11 + 1}"].value == 0.0:
                current_ws[f"{column_id}{i + 10 + 1}"].value = 0
            else:
                current_ws[f"{column_id}{i + 10 + 1}"].value = (
                        float(current_ws[f"{column_id}{i + 2 + 1}"].value) / float(
                    current_ws[f"{column_id}{i + 11 + 1}"].value))
            print(f"********** {target_worksheet_name} --含税产品-- [{target_name}]已填完 ***************")

    print("\n")
    # 根据C列单元格清空后续列
    column_values = list(current_ws["C"])

    for idx, cell in enumerate(column_values):
        if cell.value is None:
            if not any(cell.coordinate in merged_range for merged_range in current_ws.merged_cells.ranges):
                current_ws[f"{column_id}{idx + 1}"].value = None
            # current_ws[f"{column_id}{idx + 1}"].value = None
            # for row in current_ws.iter_rows(min_row=idx + 1, min_col=4, max_col=current_ws.max_column):
            #     for cell in row:
            #         cell.value = None
    # 保存并关闭工作簿
    current_wb.save(current_workbook_name + '.xlsx')
    current_wb.close()
    # target_wb.close()


if __name__ == '__main__':
    # 切换到目标文件夹
    os.chdir(r'C:\Users\EDY\Desktop')

    # 填写工作表名称
    current_workbook_name = r'2月1日经营报表'
    # 查找工作表名称
    target_workbook_name = r'2月每日数据副表'

    # 填写目标列（即需要填写日期的列名）和对应副表名称
    date_args = {
        "D": ["2.1小澳", "2.1大洋洲", "健康宝宝2.1"],
        "E": ["2.2小澳", "2.2大洋洲", "2.2健康宝宝"],
        "F": ["2.3小澳", "2.3大洋洲", "2.3健康宝宝"],
    }
    for column_name, store in date_args.items():
        # "填写工作簿": ["查找的工作簿", 开始行数, 考虑税产品id开始行数, 结尾行数]
        dict_arg = {
            "小澳产品经营报表": [store[0], 3, 147, 550],
            "大洋洲产品经营报表": [store[1], 3, 129, 670],
            "健康宝宝产品经营报表": [store[2], 3, 3, 350],
        }

        for current, target in dict_arg.items():
            # 不考虑税产品-填写操作
            v_lookup(current_workbook_name, current, target_workbook_name, target[0], column_name, target[1], target[2])
            # 考虑税产品-填写操作
            v_lookup_with_tax(current_workbook_name, current, target_workbook_name, target[0], column_name, target[2],
                              target[3])
            print("*" * 120 + "\n")
