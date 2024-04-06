#!/usr/bin/env python3
# coding:utf-8
import openpyxl
import xlsxwriter
import re
import os
import csv
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import json
import threading
import random
import pandas as pd


class Bag:
    def __init__(self):
        pass

    @staticmethod
    def web_gpu():
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
        # chrome_options.add_argument('blink-settings=imagesEnabled=false')
        options = chrome_options
        web = Chrome(options=options)
        return web

    @staticmethod
    def web_cpu(ua=r'Opera/9.80 (Windows NT 6.1; U; zh-cn) Presto/2.6.37 Version/11.00'):
        chrome_options = Options()
        chrome_options.add_argument('--user-agent={}'.format(ua))
        chrome_options.add_argument('--disable-gpu')
        return Chrome(options=chrome_options)

    @staticmethod
    def web_debug():
        chrome_options = Options()
        chrome_options.page_load_strategy = 'eager'
        chrome_options.add_argument('--blink-setting=imagesEnabled=false')
        chrome_options.add_argument('--disable-gup')
        chrome_options.add_argument('--disable-javascript')
        chrome_options.add_argument('--enable-automation')
        chrome_options.add_experimental_option('debuggerAddress', '127.0.0.1:9222')
        web = Chrome(service=Service(), options=chrome_options, chrome_options=chrome_options)
        return web

    @staticmethod
    def web_no_pic():
        chrome_options = Options()
        chrome_options.page_load_strategy = 'eager'
        chrome_options.add_argument('--disable-gup')
        chrome_options.add_argument('--disable-javascript')
        chrome_options.add_argument('--enable-automation')
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.2.13) Gecko/20101213 Opera/9.80 (Windows NT 6.1; U; zh-tw) Presto/2.7.62 Version/11.01')
        prefs = {
            'profile.default_content_setting_values': {
                'images': 2,
            }
        }
        chrome_options.add_experimental_option('prefs', prefs)
        web = Chrome(options=chrome_options)
        return web

    @staticmethod
    def web_drop_down():
        web = Bag.web_cpu()
        js = r"window.scrollTo(0, document.body.scrollHeight)"
        web.execute_script(js)
        return web

    @staticmethod
    def web_custom_ua(ua=r'Opera/9.80 (Windows NT 6.1; U; zh-cn) Presto/2.6.37 Version/11.00'):
        chrome_options = Options()
        chrome_options.add_argument('user-agent={}'.format(ua))
        chrome_options.add_argument('--disable-gpu')
        web = Chrome(options=chrome_options)
        return web

    @staticmethod
    def read_excel(path, sheet_name='Sheet1', column_num=None):
        data = []
        book = openpyxl.load_workbook(path)
        sheet = book[sheet_name]
        rows = sheet.max_row
        columns = sheet.max_column
        if bool(column_num):
            for _i in range(rows):
                data.append([str(sheet.cell(_i+1, num).value) for num in column_num])
        else:
            for _i in range(rows):
                data.append([str(sheet.cell(_i+1, num+1).value) for num in range(columns)])
        book.close()
        return data

    @staticmethod
    def save_excel(_ls, path, _sheet_name='Sheet1'):
        _book = xlsxwriter.Workbook(path, options={'strings_to_urls': False})
        _sheet = _book.add_worksheet(_sheet_name)
        for _i in range(len(_ls)):
            for __i in range(len(_ls[_i])):
                _sheet.write(_i, __i, str(_ls[_i][__i]))
        _book.close()

    @staticmethod
    def read_csv(path, encoding='gbk'):
        _data = []
        with open(path, encoding=encoding, mode='r') as cf:
            reader = csv.reader(cf)
            for _i in reader:
                _data.append(_i)
        return _data

    @staticmethod
    def save_csv(_ls, path):
        with open(path, encoding='utf8', mode='a+', newline='') as wf:
            writer = csv.writer(wf)
            writer.writerows(_ls)

    @staticmethod
    def read_json(path):
        with open(path, encoding='utf8', mode='r') as jf:
            return json.load(jf)

    @staticmethod
    def save_json(_data, path='./resp.json'):
        with open(path, encoding='utf8', mode='w') as wjf:
            json.dump(_data, wjf, ensure_ascii=False, indent=4)

    @staticmethod
    def headers():
        UA = []
        with open('ua.txt', mode='r', encoding='utf8') as file:
            for i in file.readlines():
                UA.append(i.replace('\n', ''))
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'zh-CN,zh;q=0.9',
            'cache-control': 'max-age=0',
            'cookie': 'session-id=136-1324558-3473404; ubid-main=135-6357305-1283451; i18n-prefs=USD; lc-main=en_US; av-timezone=Asia/Shanghai; x-main="HvrvNb5lO6TofumPmvlLMDDHVn1dgD6t9nuQAdbmJnJljQ5JWgsqU6K?eshz@Wx3"; at-main=Atza|IwEBIGCstlLxZtNs4mxHNTlQXw6WqZ5NHGKXRqxiVznCn1cYbJTlIfUcfJOQ5ZxFNMojXKaQSzuCBqGBYMReHYaQJf935dqR8juG2OXtWHlYus_3iba_CyzLgjtxNujiEoNWJEI_tSoNG02-K9HSPHjY7sB6adVqrLkGRlsMB0u39dSyFWttSeX6DRZmHXs3fLWRhPtuxdO43Fy4Egz9I1CWQKo7; sess-at-main="xJNuoF0w7BCCrn9Dc7BZGUMxeJ87jZ7MK42Pe8qlS6c="; sst-main=Sst1|PQH17pAJTYBNQFvYBkgbeRWsCSttm_3eHL_YHqJxuNcad51fUZLJq-sWWAFd1oy5L9sY6T3_fTNlGTU7uvC4vkduXbxwKA5XcszRGxb2QgqWxU0F3_pGgzzOz6L4fh9BNYA8TGrgj6O06BoUzuQCEo89W4a2ktNFeOGo1JOEJlLOiT4z2Qhy-e5Y1Lhyhb2CIg21MEaxPFNwr6xS6Ni15U-jmkW1CaqWx_TIAn6HBDzAgSqKo0-BlRL74cxJ-0K7T9vAvLodw1Hbl0yLQGvFKXjQekRqTtlYtfD282oYdGZC3Lk; session-id-time=2082787201l; session-token=j/tzNCHzjkdcuPyjcbVumvxWS8XRGJQJ/dmtZfsAbnFXMflF5BpYSCu3a/TiZ7gkveDlA14VOlY9ig11bO/NyIxsuVHx/u6Hln647TskeOjl9o3paklXo6GUyhgLp11zrikSErvt4njvlJ6Zc5VOCyx92qsWPyFnJjtkqmHNHErBGubBeiW79yuamtNH3GvlsFt+BDh7aRTJAbNekBfsPDeSJQhpt6lPfQC2gc9WF5LCMyyHgglwh3hXY+z0HBaz; csm-hit=tb:B38M3MDAA5DQB7EGQHD8+s-B38M3MDAA5DQB7EGQHD8|1671779406666&t:1671779406666&adb:adblk_no',
            'device-memory': '8',
            'downlink': '10',
            'dpr': '1',
            'ect': '4g',
            'rtt': '200',
            'sec-ch-device-memory': '8',
            'sec-ch-dpr': '1',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-ch-viewport-width': '1087',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'sec-fetch-user': '?1',
            'upgrade-insecure-requests': '1',
            'user-agent': '{}'.format(UA[random.randint(1, 300)])
        }
        return headers


# noinspection PyBroadException
class MyThread(threading.Thread):
    def __init__(self, func, args=()):
        super(MyThread, self).__init__()
        self.result = None
        self.func = func
        self.args = args

    def run(self):
        self.result = self.func(*self.args)

    def get_result(self):
        try:
            return self.result
        except Exception as e:
            return None


class FileFormat:
    def __init__(self):
        pass

    @staticmethod
    def remove(path, save_path, sheet_name='Sheet1', tag=''):
        df = pd.read_excel(path, sheet_name=sheet_name)
        df = df.drop_duplicates(tag)
        df.to_excel(excel_writer=save_path, index=False)


class Verify:
    def __int__(self):
        pass

